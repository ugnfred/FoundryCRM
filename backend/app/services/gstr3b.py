"""GSTR-3B computation and Excel generation."""
from decimal import Decimal
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def _d(v) -> Decimal:
    return Decimal(str(v or 0))


def _amt(v) -> float:
    return round(float(v), 2)


def compute_gstr3b(invoices: list[dict], credit_notes: list[dict], purchase_orders: list[dict]) -> dict:
    """Return GSTR-3B summary from invoice, CN, and PO data."""
    # 3.1 Outward supplies
    outward_taxable = outward_cgst = outward_sgst = outward_igst = Decimal(0)
    nil_rated = exempt = Decimal(0)

    for inv in invoices:
        if inv.get("status") in ("draft", "cancelled"):
            continue
        outward_taxable += _d(inv.get("taxable_amt", 0))
        outward_cgst += _d(inv.get("cgst", 0))
        outward_sgst += _d(inv.get("sgst", 0))
        outward_igst += _d(inv.get("igst", 0))

    # Reduce by credit notes
    cn_taxable = cn_cgst = cn_sgst = cn_igst = Decimal(0)
    for cn in credit_notes:
        if cn.get("status") == "cancelled":
            continue
        cn_taxable += _d(cn.get("taxable_amt", 0))
        cn_cgst += _d(cn.get("cgst", 0))
        cn_sgst += _d(cn.get("sgst", 0))
        cn_igst += _d(cn.get("igst", 0))

    net_taxable = outward_taxable - cn_taxable
    net_cgst = outward_cgst - cn_cgst
    net_sgst = outward_sgst - cn_sgst
    net_igst = outward_igst - cn_igst

    # 4 ITC — from purchase orders (input GST)
    itc_cgst = itc_sgst = itc_igst = Decimal(0)
    for po in purchase_orders:
        if po.get("status") == "cancelled":
            continue
        # POs store total_gst but not CGST/SGST/IGST split. Approximate 50/50 if intra.
        gst = _d(po.get("total_gst", 0))
        itc_cgst += gst / 2
        itc_sgst += gst / 2

    # 5 Tax payable (GST collected - ITC)
    tax_cgst = max(Decimal(0), net_cgst - itc_cgst)
    tax_sgst = max(Decimal(0), net_sgst - itc_sgst)
    tax_igst = max(Decimal(0), net_igst - itc_igst)

    return {
        "3_1": {
            "a_taxable": _amt(net_taxable),
            "a_cgst": _amt(net_cgst),
            "a_sgst": _amt(net_sgst),
            "a_igst": _amt(net_igst),
            "b_nil_rated": _amt(nil_rated),
            "c_exempt": _amt(exempt),
        },
        "4": {
            "itc_cgst": _amt(itc_cgst),
            "itc_sgst": _amt(itc_sgst),
            "itc_igst": _amt(itc_igst),
            "itc_total": _amt(itc_cgst + itc_sgst + itc_igst),
        },
        "5": {
            "cgst": _amt(tax_cgst),
            "sgst": _amt(tax_sgst),
            "igst": _amt(tax_igst),
            "total": _amt(tax_cgst + tax_sgst + tax_igst),
        },
        "credit_notes": {
            "taxable": _amt(cn_taxable),
            "cgst": _amt(cn_cgst),
            "sgst": _amt(cn_sgst),
            "igst": _amt(cn_igst),
        },
    }


def generate_gstr3b_excel(data: dict, period: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GSTR-3B"

    blue_fill = PatternFill("solid", fgColor="1a56db")
    hdr_font = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)

    def hdr(row, col, val):
        c = ws.cell(row, col, val)
        c.fill = blue_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="left")
        return c

    ws["A1"] = f"GSTR-3B — {period}"
    ws["A1"].font = Font(bold=True, size=14, color="1a56db")
    ws.merge_cells("A1:E1")

    r = 3
    hdr(r, 1, "Table")
    hdr(r, 2, "Description")
    hdr(r, 3, "CGST (₹)")
    hdr(r, 4, "SGST (₹)")
    hdr(r, 5, "IGST (₹)")
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    r = 4
    s31 = data["3_1"]
    rows_3_1 = [
        ("3.1(a)", "Outward taxable supplies (other than zero-rated/nil/exempt)",
         s31["a_cgst"], s31["a_sgst"], s31["a_igst"]),
        ("3.1(b)", "Nil-rated supplies", 0, 0, 0),
        ("3.1(c)", "Exempt supplies", 0, 0, 0),
        ("3.1(d)", "Zero-rated supplies", 0, 0, 0),
    ]
    for tbl, desc, cgst, sgst, igst in rows_3_1:
        ws.cell(r, 1, tbl)
        ws.cell(r, 2, desc)
        ws.cell(r, 3, cgst)
        ws.cell(r, 4, sgst)
        ws.cell(r, 5, igst)
        r += 1

    # Section 4 — ITC
    r += 1
    hdr(r, 1, "4")
    hdr(r, 2, "Eligible ITC")
    hdr(r, 3, "CGST (₹)")
    hdr(r, 4, "SGST (₹)")
    hdr(r, 5, "IGST (₹)")
    r += 1
    itc = data["4"]
    ws.cell(r, 1, "4(A)")
    ws.cell(r, 2, "ITC Available (all other ITC — from purchases)")
    ws.cell(r, 3, itc["itc_cgst"])
    ws.cell(r, 4, itc["itc_sgst"])
    ws.cell(r, 5, itc["itc_igst"])
    r += 2

    # Section 5 — Tax payable
    hdr(r, 1, "5")
    hdr(r, 2, "Net Tax Payable")
    hdr(r, 3, "CGST (₹)")
    hdr(r, 4, "SGST (₹)")
    hdr(r, 5, "IGST (₹)")
    r += 1
    s5 = data["5"]
    ws.cell(r, 1, "5(A)")
    ws.cell(r, 2, "Tax payable after ITC")
    ws.cell(r, 3, s5["cgst"])
    ws.cell(r, 4, s5["sgst"])
    ws.cell(r, 5, s5["igst"])
    ws.cell(r, 3).font = bold
    ws.cell(r, 4).font = bold
    ws.cell(r, 5).font = bold
    r += 2

    # Credit notes summary
    hdr(r, 1, "CDN")
    hdr(r, 2, "Credit Notes Issued (deducted from 3.1(a))")
    hdr(r, 3, "CGST (₹)")
    hdr(r, 4, "SGST (₹)")
    hdr(r, 5, "IGST (₹)")
    r += 1
    cn = data["credit_notes"]
    ws.cell(r, 1, "")
    ws.cell(r, 2, "Total credit notes")
    ws.cell(r, 3, cn["cgst"])
    ws.cell(r, 4, cn["sgst"])
    ws.cell(r, 5, cn["igst"])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
