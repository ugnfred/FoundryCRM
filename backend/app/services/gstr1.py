"""GSTR-1 computation and Excel generation."""
from decimal import Decimal
from datetime import date
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def _d(v) -> Decimal:
    return Decimal(str(v or 0))


def compute_gstr1(invoices: list[dict], credit_notes: list[dict]) -> dict:
    """Return structured GSTR-1 data from issued invoices and credit notes."""
    b2b = {}      # GSTIN -> list of invoice dicts
    b2cs = {}     # state_code -> aggregated amounts
    hsn_map = {}  # hsn_code -> aggregated amounts

    for inv in invoices:
        if inv.get("status") in ("draft", "cancelled"):
            continue
        company = inv.get("companies") or {}
        gstin = company.get("gstin", "").strip()
        pos = inv.get("place_of_supply", "")
        txable = _d(inv.get("taxable_amt", 0))
        cgst = _d(inv.get("cgst", 0))
        sgst = _d(inv.get("sgst", 0))
        igst = _d(inv.get("igst", 0))
        total = _d(inv.get("total", 0))

        inv_rec = {
            "inv_no": inv.get("inv_no", ""),
            "date": inv.get("date", ""),
            "place_of_supply": pos,
            "taxable_amt": txable,
            "cgst": cgst,
            "sgst": sgst,
            "igst": igst,
            "total": total,
            "customer": company.get("name", ""),
            "gstin": gstin,
        }

        # B2B: customer has a GSTIN (registered buyer)
        if gstin:
            b2b.setdefault(gstin, {"name": company.get("name", ""), "gstin": gstin, "invoices": []})
            b2b[gstin]["invoices"].append(inv_rec)
        else:
            # B2CS: unregistered buyer — aggregate per state
            key = pos or "99"
            if key not in b2cs:
                b2cs[key] = {"place_of_supply": key, "taxable_amt": Decimal(0), "cgst": Decimal(0), "sgst": Decimal(0), "igst": Decimal(0), "total": Decimal(0)}
            b2cs[key]["taxable_amt"] += txable
            b2cs[key]["cgst"] += cgst
            b2cs[key]["sgst"] += sgst
            b2cs[key]["igst"] += igst
            b2cs[key]["total"] += total

        # HSN summary
        for item in inv.get("invoice_items", []):
            hsn = item.get("hsn_code", "MISC") or "MISC"
            desc = item.get("description", "")
            uom = item.get("uom", "NOS")
            gst_rate = _d(item.get("gst_rate", 0))
            qty = _d(item.get("qty", 0))
            item_taxable = _d(item.get("qty", 0)) * _d(item.get("rate", 0))
            item_igst = _d(item.get("igst_amt", 0))
            item_cgst = _d(item.get("cgst_amt", 0))
            item_sgst = _d(item.get("sgst_amt", 0))

            if hsn not in hsn_map:
                hsn_map[hsn] = {"hsn_code": hsn, "description": desc, "uom": uom, "gst_rate": gst_rate,
                                 "qty": Decimal(0), "taxable_amt": Decimal(0), "igst": Decimal(0), "cgst": Decimal(0), "sgst": Decimal(0)}
            hsn_map[hsn]["qty"] += qty
            hsn_map[hsn]["taxable_amt"] += item_taxable
            hsn_map[hsn]["igst"] += item_igst
            hsn_map[hsn]["cgst"] += item_cgst
            hsn_map[hsn]["sgst"] += item_sgst

    # Credit notes — CDNR (against registered) and CDNS (against unregistered)
    cdnr = {}
    cdns = []
    for cn in credit_notes:
        if cn.get("status") == "cancelled":
            continue
        company = cn.get("companies") or {}
        gstin = company.get("gstin", "").strip()
        cn_rec = {
            "cn_no": cn.get("cn_no", ""),
            "date": cn.get("date", ""),
            "reason": cn.get("reason", ""),
            "taxable_amt": _d(cn.get("taxable_amt", 0)),
            "cgst": _d(cn.get("cgst", 0)),
            "sgst": _d(cn.get("sgst", 0)),
            "igst": _d(cn.get("igst", 0)),
            "total": _d(cn.get("total", 0)),
            "customer": company.get("name", ""),
            "linked_inv": (cn.get("invoices") or {}).get("inv_no", ""),
        }
        if gstin:
            cdnr.setdefault(gstin, {"name": company.get("name", ""), "gstin": gstin, "notes": []})
            cdnr[gstin]["notes"].append(cn_rec)
        else:
            cdns.append(cn_rec)

    return {
        "b2b": list(b2b.values()),
        "b2cs": list(b2cs.values()),
        "cdnr": list(cdnr.values()),
        "cdns": cdns,
        "hsn": list(hsn_map.values()),
    }


def _header_row(ws, row: int, values: list, fill_hex: str = "1a56db"):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color="FFFFFF")
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _border_thin():
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)


def _amt(v) -> float:
    return round(float(v), 2)


def generate_gstr1_excel(data: dict, period: str) -> bytes:
    """Return GSTR-1 Excel workbook as bytes."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: B2B ──────────────────────────────────────────
    ws = wb.active
    ws.title = "B2B"
    _header_row(ws, 1, ["GSTIN", "Customer", "Invoice No", "Date", "Place of Supply",
                         "Taxable (₹)", "CGST (₹)", "SGST (₹)", "IGST (₹)", "Total (₹)"])
    r = 2
    for entry in data["b2b"]:
        for inv in entry["invoices"]:
            ws.cell(r, 1, entry.get("name", "") + "\n" + inv["gstin"]).alignment = Alignment(wrap_text=True)
            ws.cell(r, 1, inv["gstin"])
            ws.cell(r, 2, entry["name"])
            ws.cell(r, 3, inv["inv_no"])
            ws.cell(r, 4, str(inv["date"]))
            ws.cell(r, 5, inv["place_of_supply"])
            ws.cell(r, 6, _amt(inv["taxable_amt"]))
            ws.cell(r, 7, _amt(inv["cgst"]))
            ws.cell(r, 8, _amt(inv["sgst"]))
            ws.cell(r, 9, _amt(inv["igst"]))
            ws.cell(r, 10, _amt(inv["total"]))
            r += 1
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14

    # ── Sheet 2: B2CS ─────────────────────────────────────────
    ws2 = wb.create_sheet("B2CS")
    _header_row(ws2, 1, ["Place of Supply", "Taxable (₹)", "CGST (₹)", "SGST (₹)", "IGST (₹)", "Total (₹)"], "059669")
    r = 2
    for row in data["b2cs"]:
        ws2.cell(r, 1, row["place_of_supply"])
        ws2.cell(r, 2, _amt(row["taxable_amt"]))
        ws2.cell(r, 3, _amt(row["cgst"]))
        ws2.cell(r, 4, _amt(row["sgst"]))
        ws2.cell(r, 5, _amt(row["igst"]))
        ws2.cell(r, 6, _amt(row["total"]))
        r += 1

    # ── Sheet 3: CDNR (Credit Notes – Registered) ─────────────
    ws3 = wb.create_sheet("CDNR")
    _header_row(ws3, 1, ["GSTIN", "Customer", "CN No", "Date", "Linked Invoice",
                          "Reason", "Taxable (₹)", "CGST (₹)", "SGST (₹)", "IGST (₹)", "Total (₹)"], "dc2626")
    r = 2
    for entry in data["cdnr"]:
        for cn in entry["notes"]:
            ws3.cell(r, 1, entry["name"])
            ws3.cell(r, 2, cn["customer"])
            ws3.cell(r, 3, cn["cn_no"])
            ws3.cell(r, 4, str(cn["date"]))
            ws3.cell(r, 5, cn["linked_inv"])
            ws3.cell(r, 6, cn["reason"])
            ws3.cell(r, 7, _amt(cn["taxable_amt"]))
            ws3.cell(r, 8, _amt(cn["cgst"]))
            ws3.cell(r, 9, _amt(cn["sgst"]))
            ws3.cell(r, 10, _amt(cn["igst"]))
            ws3.cell(r, 11, _amt(cn["total"]))
            r += 1

    # ── Sheet 4: HSN Summary ──────────────────────────────────
    ws4 = wb.create_sheet("HSN Summary")
    _header_row(ws4, 1, ["HSN Code", "Description", "UOM", "GST Rate (%)", "Qty",
                          "Taxable (₹)", "CGST (₹)", "SGST (₹)", "IGST (₹)"], "7c3aed")
    r = 2
    for row in data["hsn"]:
        ws4.cell(r, 1, row["hsn_code"])
        ws4.cell(r, 2, row["description"])
        ws4.cell(r, 3, row["uom"])
        ws4.cell(r, 4, _amt(row["gst_rate"]))
        ws4.cell(r, 5, _amt(row["qty"]))
        ws4.cell(r, 6, _amt(row["taxable_amt"]))
        ws4.cell(r, 7, _amt(row["cgst"]))
        ws4.cell(r, 8, _amt(row["sgst"]))
        ws4.cell(r, 9, _amt(row["igst"]))
        r += 1
    ws4.column_dimensions["B"].width = 30

    # ── Sheet 5: Summary ──────────────────────────────────────
    ws5 = wb.create_sheet("Summary")
    ws5.title = "Summary"
    ws5["A1"] = f"GSTR-1 — {period}"
    ws5["A1"].font = Font(bold=True, size=14)

    totals_b2b = _sum_b2b(data["b2b"])
    totals_b2cs = _sum_b2cs(data["b2cs"])
    totals_cdnr = _sum_cdnr(data["cdnr"])

    summary_rows = [
        ("", "Description", "Taxable (₹)", "CGST (₹)", "SGST (₹)", "IGST (₹)", "Total (₹)"),
        ("3A", "B2B Supplies (Registered)", totals_b2b[0], totals_b2b[1], totals_b2b[2], totals_b2b[3], totals_b2b[4]),
        ("3B", "B2CS Supplies (Unregistered)", totals_b2cs[0], totals_b2cs[1], totals_b2cs[2], totals_b2cs[3], totals_b2cs[4]),
        ("4", "Credit Notes (Registered)", -totals_cdnr[0], -totals_cdnr[1], -totals_cdnr[2], -totals_cdnr[3], -totals_cdnr[4]),
    ]
    for i, row_data in enumerate(summary_rows, 3):
        for col, val in enumerate(row_data, 1):
            ws5.cell(i, col, val)
        if i == 3:
            for col in range(1, 8):
                ws5.cell(i, col).font = Font(bold=True)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sum_b2b(b2b_list):
    t = cgst = sgst = igst = total = Decimal(0)
    for entry in b2b_list:
        for inv in entry["invoices"]:
            t += inv["taxable_amt"]
            cgst += inv["cgst"]
            sgst += inv["sgst"]
            igst += inv["igst"]
            total += inv["total"]
    return _amt(t), _amt(cgst), _amt(sgst), _amt(igst), _amt(total)


def _sum_b2cs(b2cs_list):
    t = cgst = sgst = igst = total = Decimal(0)
    for row in b2cs_list:
        t += row["taxable_amt"]
        cgst += row["cgst"]
        sgst += row["sgst"]
        igst += row["igst"]
        total += row["total"]
    return _amt(t), _amt(cgst), _amt(sgst), _amt(igst), _amt(total)


def _sum_cdnr(cdnr_list):
    t = cgst = sgst = igst = total = Decimal(0)
    for entry in cdnr_list:
        for cn in entry["notes"]:
            t += cn["taxable_amt"]
            cgst += cn["cgst"]
            sgst += cn["sgst"]
            igst += cn["igst"]
            total += cn["total"]
    return _amt(t), _amt(cgst), _amt(sgst), _amt(igst), _amt(total)
