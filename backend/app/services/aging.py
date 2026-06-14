"""Receivables and Payables aging computation and Excel generation."""
from decimal import Decimal
from datetime import date
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def _d(v) -> Decimal:
    return Decimal(str(v or 0))


def _amt(v) -> float:
    return round(float(v), 2)


BUCKETS = [(0, 30), (31, 60), (61, 90), (91, 120), (121, None)]
BUCKET_LABELS = ["0-30 days", "31-60 days", "61-90 days", "91-120 days", "120+ days"]


def _bucket_index(days: int) -> int:
    for i, (lo, hi) in enumerate(BUCKETS):
        if hi is None or days <= hi:
            return i
    return len(BUCKETS) - 1


def compute_receivables_aging(invoices: list[dict], as_of: date) -> list[dict]:
    """Return per-customer aging rows with bucket amounts."""
    customers = {}
    for inv in invoices:
        if inv.get("status") in ("draft", "cancelled", "paid"):
            continue
        balance = _d(inv.get("balance_due", 0))
        if balance <= 0:
            continue
        company = inv.get("companies") or {}
        cust_id = inv.get("company_id", "")
        cust_name = company.get("name", "Unknown")
        inv_date = inv.get("due_date") or inv.get("date") or str(as_of)
        try:
            inv_d = date.fromisoformat(str(inv_date))
        except ValueError:
            inv_d = as_of
        days = (as_of - inv_d).days
        days = max(0, days)
        bi = _bucket_index(days)

        if cust_id not in customers:
            customers[cust_id] = {
                "customer_id": cust_id,
                "customer_name": cust_name,
                "gstin": company.get("gstin", ""),
                "buckets": [Decimal(0)] * len(BUCKETS),
                "total": Decimal(0),
            }
        customers[cust_id]["buckets"][bi] += balance
        customers[cust_id]["total"] += balance

    result = []
    for row in sorted(customers.values(), key=lambda x: -float(x["total"])):
        result.append({
            "customer_name": row["customer_name"],
            "gstin": row["gstin"],
            **{BUCKET_LABELS[i]: _amt(row["buckets"][i]) for i in range(len(BUCKETS))},
            "total": _amt(row["total"]),
        })
    return result


def compute_payables_aging(purchase_orders: list[dict], as_of: date) -> list[dict]:
    """Return per-supplier aging rows with bucket amounts."""
    suppliers = {}
    for po in purchase_orders:
        if po.get("status") in ("cancelled", "closed"):
            continue
        # POs don't have a balance_due — use total as outstanding
        total = _d(po.get("total", 0))
        if total <= 0:
            continue
        company = po.get("companies") or {}
        sup_id = po.get("company_id", "")
        sup_name = company.get("name", "Unknown")
        po_date = po.get("delivery_date") or po.get("date") or str(as_of)
        try:
            po_d = date.fromisoformat(str(po_date))
        except ValueError:
            po_d = as_of
        days = max(0, (as_of - po_d).days)
        bi = _bucket_index(days)

        if sup_id not in suppliers:
            suppliers[sup_id] = {
                "supplier_id": sup_id,
                "supplier_name": sup_name,
                "gstin": company.get("gstin", ""),
                "buckets": [Decimal(0)] * len(BUCKETS),
                "total": Decimal(0),
            }
        suppliers[sup_id]["buckets"][bi] += total
        suppliers[sup_id]["total"] += total

    result = []
    for row in sorted(suppliers.values(), key=lambda x: -float(x["total"])):
        result.append({
            "supplier_name": row["supplier_name"],
            "gstin": row["gstin"],
            **{BUCKET_LABELS[i]: _amt(row["buckets"][i]) for i in range(len(BUCKETS))},
            "total": _amt(row["total"]),
        })
    return result


def generate_aging_excel(rows: list[dict], report_title: str, name_col: str, as_of: date) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_title[:30]

    blue = PatternFill("solid", fgColor="1a56db")
    hdr_font = Font(bold=True, color="FFFFFF")
    red_fill = PatternFill("solid", fgColor="fee2e2")
    amber_fill = PatternFill("solid", fgColor="fef3c7")

    ws["A1"] = f"{report_title} — As of {as_of}"
    ws["A1"].font = Font(bold=True, size=13, color="1a56db")

    headers = [name_col, "GSTIN"] + BUCKET_LABELS + ["Total (₹)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(3, col, h)
        cell.fill = blue
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    for col_letter in ["C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col_letter].width = 15

    r = 4
    for row in rows:
        ws.cell(r, 1, row.get(name_col, ""))
        ws.cell(r, 2, row.get("gstin", ""))
        for i, label in enumerate(BUCKET_LABELS):
            amt = row.get(label, 0)
            c = ws.cell(r, 3 + i, amt)
            if amt > 0:
                if i >= 3:
                    c.fill = red_fill
                elif i == 2:
                    c.fill = amber_fill
        ws.cell(r, 3 + len(BUCKET_LABELS), row.get("total", 0)).font = Font(bold=True)
        r += 1

    # Totals row
    if rows:
        ws.cell(r, 1, "TOTAL").font = Font(bold=True)
        for i in range(len(BUCKET_LABELS)):
            col_total = sum(row.get(BUCKET_LABELS[i], 0) for row in rows)
            ws.cell(r, 3 + i, round(col_total, 2)).font = Font(bold=True)
        grand = sum(row.get("total", 0) for row in rows)
        ws.cell(r, 3 + len(BUCKET_LABELS), round(grand, 2)).font = Font(bold=True)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
